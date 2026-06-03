import asyncio
import aiohttp
import logging
import os
import sys
import locale
import re
from dataclasses import dataclass
from itertools import chain
from typing import Any, Dict, List, Optional

# openpyxl para generar el Excel de forma nativa en Linux (Streamlit Cloud)
from openpyxl import Workbook
from openpyxl.styles import PatternFill

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

try:
    locale.setlocale(locale.LC_NUMERIC, '')
except Exception:
    pass 

BASE_URLS = {
    'subestaciones': 'https://api-infotecnica.coordinador.cl/v1/subestaciones',
    'interruptores': 'https://api-infotecnica.coordinador.cl/v1/interruptores',
    'transformadores_2d': 'https://api-infotecnica.coordinador.cl/v1/transformadores-2d',
    'transformadores_3d': 'https://api-infotecnica.coordinador.cl/v1/transformadores-3d',
    'lineas': 'https://api-infotecnica.coordinador.cl/v1/lineas',
    'tramos': 'https://api-infotecnica.coordinador.cl/v1/tramos',
    'secciones_tramos': 'https://api-infotecnica.coordinador.cl/v1/secciones-tramos',
    'desconectadores': 'https://api-infotecnica.coordinador.cl/v1/desconectadores',
    'transformadores_corriente': 'https://api-infotecnica.coordinador.cl/v1/transformadores-corrientes',
    'trampas_ondas': 'https://api-infotecnica.coordinador.cl/v1/trampas-ondas',
    'unidades_generadoras': 'https://api-infotecnica.coordinador.cl/v1/unidades-generadoras',
    'panos': 'https://api-infotecnica.coordinador.cl/v1/panos',
}

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://infotecnica.coordinador.cl/",
    "Origin": "https://infotecnica.coordinador.cl",
    "Connection": "keep-alive",
}

@dataclass
class EquipoDatosTecnicos:
    estados_certificacion: Dict[str, str]

@dataclass
class InterruptorDatosTecnicos(EquipoDatosTecnicos):
    tension_nominal: Optional[str] = None
    corriente_nominal: Optional[str] = None
    capacidad_simetrica: Optional[str] = None
    fabricante: Optional[str] = None
    capacidad_asimetrica: Optional[str] = None
    capacidad_cierre: Optional[str] = None
    modelo: Optional[str] = None

@dataclass
class TransformadorDatosTecnicos(EquipoDatosTecnicos):
    tension_at: Optional[str] = None
    tension_mt: Optional[str] = None
    capacidad_at: Optional[str] = None
    capacidad_at_onaf1: Optional[str] = None
    capacidad_at_onaf2: Optional[str] = None
    imp_sec_pos: Optional[str] = None
    pot_sec_pos: Optional[str] = None
    imp_sec_cero: Optional[str] = None
    pot_sec_cero: Optional[str] = None

@dataclass
class SeccionTramoDatosTecnicos(EquipoDatosTecnicos):
    tension_nominal: Optional[str] = None
    longitud_conductor: Optional[str] = None
    resistencia_sec_pos: Optional[str] = None
    reactancia_sec_pos: Optional[str] = None
    susceptancia_sec_pos: Optional[str] = None
    resistencia_sec_cero: Optional[str] = None
    reactancia_sec_cero: Optional[str] = None
    susceptancia_sec_cero: Optional[str] = None
    limites_termicos: Dict[str, Any] = None

@dataclass
class Equipo:
    tipo: str
    id_equipo: int
    nombre_equipo: str
    subestacion_nombre: Optional[str]
    propietario_nombre: Optional[str]
    pano_coordinado_nombre: Optional[str]
    datos_tecnicos: EquipoDatosTecnicos

@dataclass
class Interruptor(Equipo):
    datos_tecnicos: InterruptorDatosTecnicos

@dataclass
class Transformador(Equipo):
    datos_tecnicos: TransformadorDatosTecnicos

@dataclass
class SeccionTramo(Equipo):
    id_linea: int
    nombre_linea: str
    id_tramo: int
    nombre_tramo: str
    extremo1: str
    extremo2: str
    datos_tecnicos: SeccionTramoDatosTecnicos

class ApiClientFactory:
    @staticmethod
    def create_client(session: aiohttp.ClientSession) -> 'ApiClient':
        return ApiClient(session)

class ApiClient:
    def __init__(self, session: aiohttp.ClientSession):
        self.session = session
        self.nombre_subestacion_cache = {}
        self.datos_tecnicos_cache = {}

    async def hacer_solicitud(self, url: str) -> Optional[Dict[str, Any]]:
        try:
            async with self.session.get(url, headers=HEADERS) as response:
                response.raise_for_status()
                return await response.json()
        except Exception as e:
            logging.error(f"Error al realizar solicitud a {url}: {e}")
            return None

    async def buscar_nemotecnico_pano_por_extremo_tramo(self, nombre_extremo: str) -> Optional[str]:
        if not nombre_extremo: return None
        if nombre_extremo.startswith("S/E "):
            nombre_extremo = nombre_extremo[4:]
        url = f"{BASE_URLS['panos']}?nombre__icontains={nombre_extremo}"
        datos = await self.hacer_solicitud(url)
        if datos:
            for pano in datos:
                if pano.get('nombre', '').lower() == nombre_extremo.lower():
                    return pano.get('nemotecnico', '')
            return datos[0].get('nemotecnico', '')
        return None
        
    async def buscar_equipos_por_nemotecnico(self, nemotecnico_pano: str, tipo_equipo: str) -> List[Dict[str, Any]]:
        url = f"{BASE_URLS[tipo_equipo]}/?search={nemotecnico_pano}"
        datos = await self.hacer_solicitud(url)
        return datos if datos else []

    async def obtener_datos_tecnicos(self, id_equipo: int, tipo_equipo: str) -> Optional[EquipoDatosTecnicos]:
        cache_key = (id_equipo, tipo_equipo)
        if cache_key in self.datos_tecnicos_cache:
            return self.datos_tecnicos_cache[cache_key]

        if tipo_equipo == 'secciones_tramos':
            endpoint = f"{BASE_URLS['secciones_tramos']}/{id_equipo}/fichas-tecnicas/general/"
            datos = await self.hacer_solicitud(endpoint)
            limites_termicos_endpoint = f"{BASE_URLS['secciones_tramos']}/{id_equipo}/fichas-tecnicas/limites-termicos/"
            limites_termicos = await self.hacer_solicitud(limites_termicos_endpoint)
            result = None
            if datos:
                estados_certificacion = {key: value.get('estado_certificacion_nombre', '') for key, value in datos.items()}
                
                # REPARADO: Se mapean los IDs de la jerarquía que arreglaste arriba
                result = SeccionTramoDatosTecnicos(
                    tension_nominal=datos.get('5895', {}).get('valor_texto', '').replace(",", "."),
                    longitud_conductor=datos.get('1005', {}).get('valor_texto', '').replace(",", "."),
                    resistencia_sec_pos=datos.get('1008', {}).get('valor_texto', '').replace(",", "."),  
                    reactancia_sec_pos=datos.get('1009', {}).get('valor_texto', '').replace(",", "."),
                    susceptancia_sec_pos=datos.get('1010', {}).get('valor_texto', '').replace(",", "."),
                    resistencia_sec_cero=datos.get('1012', {}).get('valor_texto', '').replace(",", "."),  
                    reactancia_sec_cero=datos.get('1013', {}).get('valor_texto', '').replace(",", "."),
                    susceptancia_sec_cero=datos.get('1014', {}).get('valor_texto', '').replace(",", "."),
                    limites_termicos=limites_termicos,
                    estados_certificacion=estados_certificacion
                )
        else:
            endpoint = f"{BASE_URLS[tipo_equipo]}/{id_equipo}/fichas-tecnicas/general/"
            datos = await self.hacer_solicitud(endpoint)
            result = None
            if datos:
                estados_certificacion = {key: value.get('estado_certificacion_nombre', '') for key, value in datos.items()}
                if tipo_equipo == 'interruptores':
                    result = InterruptorDatosTecnicos(
                        tension_nominal=datos.get('6018', {}).get('valor_texto', '').replace(",", "."),
                        corriente_nominal=datos.get('6019', {}).get('valor_texto', '').replace(",", "."),
                        capacidad_simetrica=datos.get('326', {}).get('valor_texto', '').replace(",", "."),
                        fabricante=datos.get('6023', {}).get('valor_texto', ''),
                        capacidad_asimetrica=datos.get('327', {}).get('valor_texto', '').replace(",", "."),
                        capacidad_cierre=datos.get('328', {}).get('valor_texto', '').replace(",", "."),
                        modelo=datos.get('6022', {}).get('valor_texto', ''),
                        estados_certificacion=estados_certificacion
                    )
                elif tipo_equipo == 'transformadores_2d':
                    result = TransformadorDatosTecnicos(
                        tension_at=datos.get('132', {}).get('valor_texto', '').replace(",", "."),
                        tension_mt=datos.get('133', {}).get('valor_texto', '').replace(",", "."),
                        capacidad_at=datos.get('129', {}).get('valor_texto', '').replace(",", "."),
                        capacidad_at_onaf1=datos.get('130', {}).get('valor_texto', '').replace(",", "."),
                        capacidad_at_onaf2=datos.get('1917', {}).get('valor_texto', '').replace(",", "."),
                        imp_sec_pos=datos.get('136', {}).get('valor_texto', '').replace(",", "."),
                        pot_sec_pos=datos.get('137', {}).get('valor_texto', '').replace(",", "."),
                        imp_sec_cero=datos.get('6503', {}).get('valor_texto', '').replace(",", "."),
                        pot_sec_cero=datos.get('6504', {}).get('valor_texto', '').replace(",", "."),
                        estados_certificacion=estados_certificacion
                    )
        self.datos_tecnicos_cache[cache_key] = result
        return result

    async def obtener_datos_tramo(self, id_tramo: int) -> Dict[str, Any]:
        url = f"{BASE_URLS['tramos']}/{id_tramo}/"
        datos_tramo = await self.hacer_solicitud(url)
        if datos_tramo:
            return {
                'id_tramo': id_tramo,
                'nombre_tramo': datos_tramo.get('nombre', ''),
                'extremo1': self.limpiar_extremos(datos_tramo.get('extremo1_descripcion', '')),
                'extremo2': self.limpiar_extremos(datos_tramo.get('extremo2_descripcion', '')),
            }
        return {'id_tramo': id_tramo, 'nombre_tramo': '', 'extremo1': '', 'extremo2': ''}

    def limpiar_extremos(self, texto: str) -> str:
        if texto is None: return ''
        patrones = [r'^Paño\s*:\s*', r'^Tap\s*:\s*']
        for patron in patrones:
            texto = re.sub(patron, '', texto, flags=re.IGNORECASE)
        return texto.strip()

async def procesar_interruptor(id_interruptor: int, api_client: ApiClient) -> Optional[Interruptor]:
    url_interruptor = f"{BASE_URLS['interruptores']}/{id_interruptor}"
    datos_interruptor = await api_client.hacer_solicitud(url_interruptor)
    if not datos_interruptor: return None
    datos_tecnicos = await api_client.obtener_datos_tecnicos(id_interruptor, 'interruptores')
    return Interruptor('Interruptor', id_interruptor, datos_interruptor.get('nombre', ''), datos_interruptor.get('subestacion_nombre', ''), datos_interruptor.get('propietario_nombre', ''), datos_interruptor.get('pano_nombre', ''), datos_tecnicos)

async def procesar_transformador(id_transformador: int, api_client: ApiClient) -> Optional[Transformador]:
    url_transformador = f"{BASE_URLS['transformadores_2d']}/{id_transformador}"
    datos_transformador = await api_client.hacer_solicitud(url_transformador)
    if not datos_transformador: return None
    datos_tecnicos = await api_client.obtener_datos_tecnicos(id_transformador, 'transformadores_2d')
    return Transformador('Transformador 2D', id_transformador, datos_transformador.get('nombre', ''), datos_transformador.get('subestacion_nombre', ''), datos_transformador.get('propietario_nombre', ''), datos_transformador.get('coordinado_nombre', ''), datos_tecnicos)

async def procesar_seccion_tramo(id_seccion_tramo: int, api_client: ApiClient) -> Optional[SeccionTramo]:
    url_seccion_tramo = f"{BASE_URLS['secciones_tramos']}/{id_seccion_tramo}/"
    datos_seccion_tramo = await api_client.hacer_solicitud(url_seccion_tramo)
    if not datos_seccion_tramo: return None
    datos_tecnicos = await api_client.obtener_datos_tecnicos(id_seccion_tramo, 'secciones_tramos')
    tramo_info = await api_client.obtener_datos_tramo(datos_seccion_tramo.get('id_tramo', 0))
    return SeccionTramo(
        tipo='Sección Tramo', id_equipo=id_seccion_tramo, nombre_equipo=datos_seccion_tramo.get('nombre', ''),
        subestacion_nombre=None, propietario_nombre=datos_seccion_tramo.get('propietario_nombre', ''), pano_coordinado_nombre=None,
        id_linea=datos_seccion_tramo.get('id_linea', 0), nombre_linea=datos_seccion_tramo.get('linea_nombre', ''),
        id_tramo=tramo_info.get('id_tramo', 0), nombre_tramo=tramo_info.get('nombre_tramo', ''),
        extremo1=tramo_info.get('extremo1', ''), extremo2=tramo_info.get('extremo2', ''), datos_tecnicos=datos_tecnicos
    )

def aplicar_color_openpyxl(cell, estado_certificacion: str):
    colores_hex = {'Validado': 'CCFFCC', 'Rechazado': 'F4B084', 'En Uso': 'FFE699'}
    if estado_certificacion in colores_hex:
        cell.fill = PatternFill(start_color=colores_hex[estado_certificacion], end_color=colores_hex[estado_certificacion], fill_type='solid')

def crear_archivo_excel(datos_im, datos_t2d, datos_st) -> str:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    if datos_im:
        ws = wb.create_sheet(title="IM")
        ws.append(['ID Interruptor', 'Nombre Interruptor', 'Subestación', 'Propietario', 'Paño', 'Tensión Nominal [kV]', 'Corriente Nominal [A]'])
        for item in datos_im:
            dt = item.datos_tecnicos
            ws.append([item.id_equipo, item.nombre_equipo, item.subestacion_nombre, item.propietario_nombre, item.pano_coordinado_nombre, dt.tension_nominal if dt else '', dt.corriente_nominal if dt else ''])
            if dt:
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=6), dt.estados_certificacion.get('6018', ''))
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=7), dt.estados_certificacion.get('6019', ''))

    if datos_t2d:
        ws = wb.create_sheet(title="T2D")
        ws.append(['ID Transformador', 'Nombre Transformador', 'Subestación', 'Propietario', 'Coordinado', 'Tensión nominal AT [kV]', 'Tensión nominal MT [kV]'])
        for item in datos_t2d:
            dt = item.datos_tecnicos
            ws.append([item.id_equipo, item.nombre_equipo, item.subestacion_nombre, item.propietario_nombre, item.pano_coordinado_nombre, dt.tension_at if dt else '', dt.tension_mt if dt else ''])
            if dt:
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=6), dt.estados_certificacion.get('132', ''))
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=7), dt.estados_certificacion.get('133', ''))

    if datos_st:
        ws = wb.create_sheet(title="Secciones Tramos")
        # ENCABEZADOS CORREGIDOS: Se añaden las 5 columnas para límites térmicos (15°C a 35°C)
        ws.append([
            'ID Línea', 'Nombre Línea', 'ID Tramo', 'Nombre Tramo', 'ID Sección Tramo', 'Nombre Sección Tramo', 
            'Propietario', 'Extremo 1', 'Extremo 2', 'Tensión Nominal [kV]', 'Longitud conductor [km]',
            'R1 [Ohm]', 'X1 [Ohm]', 'B1 [uS]', 'R0 [Ohm]', 'X0 [Ohm]', 'B0 [uS]',
            'Lim. Térmico 15°C', 'Lim. Térmico 20°C', 'Lim. Térmico 25°C', 'Lim. Térmico 30°C', 'Lim. Térmico 35°C'
        ])
        for item in datos_st:
            dt = item.datos_tecnicos
            
            # Procesar el sub-diccionario de límites térmicos de forma segura para openpyxl
            lim_15, lim_20, lim_25, lim_30, lim_35 = '', '', '', '', ''
            est_15, est_20, est_25, est_30, est_35 = '', '', '', '', ''
            
            if dt and dt.limites_termicos:
                lt = dt.limites_termicos
                lim_15 = lt.get('1568', {}).get('valor_texto', '').replace(",", ".")
                lim_20 = lt.get('1569', {}).get('valor_texto', '').replace(",", ".")
                lim_25 = lt.get('1571', {}).get('valor_texto', '').replace(",", ".")
                lim_30 = lt.get('1573', {}).get('valor_texto', '').replace(",", ".")
                lim_35 = lt.get('1575', {}).get('valor_texto', '').replace(",", ".")
                
                est_15 = lt.get('1568', {}).get('estado_certificacion_nombre', '')
                est_20 = lt.get('1569', {}).get('estado_certificacion_nombre', '')
                est_25 = lt.get('1571', {}).get('estado_certificacion_nombre', '')
                est_30 = lt.get('1573', {}).get('estado_certificacion_nombre', '')
                est_35 = lt.get('1575', {}).get('estado_certificacion_nombre', '')

            ws.append([
                item.id_linea, item.nombre_linea, item.id_tramo, item.nombre_tramo, item.id_equipo, item.nombre_equipo, 
                item.propietario_nombre, item.extremo1, item.extremo2, 
                dt.tension_nominal if dt else '', 
                dt.longitud_conductor if dt else '',
                dt.resistencia_sec_pos if dt else '',
                dt.reactancia_sec_pos if dt else '',
                dt.susceptancia_sec_pos if dt else '',
                dt.resistencia_sec_cero if dt else '',
                dt.reactancia_sec_cero if dt else '',
                dt.susceptancia_sec_cero if dt else '',
                lim_15, lim_20, lim_25, lim_30, lim_35
            ])
            
            if dt:
                # Colores de parámetros base
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=10), dt.estados_certificacion.get('5895', ''))
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=11), dt.estados_certificacion.get('1005', ''))
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=12), dt.estados_certificacion.get('1008', '')) 
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=13), dt.estados_certificacion.get('1009', '')) 
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=14), dt.estados_certificacion.get('1010', '')) 
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=15), dt.estados_certificacion.get('1012', '')) 
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=16), dt.estados_certificacion.get('1013', '')) 
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=17), dt.estados_certificacion.get('1014', ''))
                
                # NUEVO: Colores de certificación para las celdas de límites térmicos (Columnas 18 a 22)
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=18), est_15)
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=19), est_20)
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=20), est_25)
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=21), est_30)
                aplicar_color_openpyxl(ws.cell(row=ws.max_row, column=22), est_35)

    filepath = os.path.abspath(os.path.join(os.getcwd(), "equipos_datos.xlsx"))
    wb.save(filepath)
    return filepath

async def ejecutar_extraccion_motor(list_ids: List[int], es_modo_tramo: bool) -> str:
    async with aiohttp.ClientSession() as session:
        api_client = ApiClientFactory.create_client(session)
        r_int, r_t2d, r_st = [], [], []

        if es_modo_tramo:
            for t_id in list_ids:
                seccion = await procesar_seccion_tramo(t_id, api_client)
                if seccion:
                    r_st.append(seccion)
                    for extremo in [seccion.extremo1, seccion.extremo2]:
                        if extremo and isinstance(extremo, str):
                            nemotecnico = await api_client.buscar_nemotecnico_pano_por_extremo_tramo(extremo)
                            if nemotecnico:
                                int_data = await api_client.buscar_equipos_por_nemotecnico(nemotecnico, 'interruptores')
                                for eq in int_data:
                                    r_int.append(await procesar_interruptor(eq['id'], api_client))
        else:
            for eq_id in list_ids:
                interruptor = await procesar_interruptor(eq_id, api_client)
                if interruptor and interruptor.datos_tecnicos:
                    r_int.append(interruptor)
                    continue
                
                transformador = await procesar_transformador(eq_id, api_client)
                if transformador and transformador.datos_tecnicos:
                    r_t2d.append(transformador)
                    continue

        r_int = [x for x in r_int if x]; r_t2d = [x for x in r_t2d if x]; r_st = [x for x in r_st if x]
        if not any([r_int, r_t2d, r_st]):
            raise ValueError("No se encontraron registros válidos para esos IDs en el modo seleccionado.")
        return crear_archivo_excel(r_int, r_t2d, r_st)
