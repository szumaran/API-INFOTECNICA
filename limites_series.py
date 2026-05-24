import asyncio
import aiohttp
import logging
import os
import re
from typing import Any, Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

BASE_URLS = {
    'interruptores': 'https://api-infotecnica.coordinador.cl/v1/interruptores',
    'transformadores_2d': 'https://api-infotecnica.coordinador.cl/v1/transformadores-2d',
    'transformadores_3d': 'https://api-infotecnica.coordinador.cl/v1/transformadores-3d',
    'secciones_tramos': 'https://api-infotecnica.coordinador.cl/v1/secciones-tramos',
    'tramos': 'https://api-infotecnica.coordinador.cl/v1/tramos',
    'desconectadores': 'https://api-infotecnica.coordinador.cl/v1/desconectadores',
    'transformadores_corriente': 'https://api-infotecnica.coordinador.cl/v1/transformadores-corrientes',
    'trampas_ondas': 'https://api-infotecnica.coordinador.cl/v1/trampas-ondas',
    'panos': 'https://api-infotecnica.coordinador.cl/v1/panos',
}

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://infotecnica.coordinador.cl/",
    "Origin": "https://infotecnica.coordinador.cl",
    "Connection": "keep-alive",
}

def limpiar_valor_float(texto_valor: Any) -> float:
    try:
        if not texto_valor: return float('inf')
        num_str = re.sub(r'[^\d,\.]', '', str(texto_valor)).replace(",", ".")
        return float(num_str) if num_str else float('inf')
    except Exception:
        return float('inf')

async def buscar_limites_series_motor(list_ids: List[int], es_modo_tramo: bool) -> str:
    async with aiohttp.ClientSession() as session:
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Análisis de Elementos en Serie"
        
        filas_reporte = []
        max_equipos_detectados = 0

        for eq_id in list_ids:
            panos_a_procesar = [] # Guardaremos tuplas de (id_del_paño, nemotecnico_del_paño)
            subestacion = 'Desconocida'

            if es_modo_tramo:
                # ==============================================================
                # MODO TRAMO: Buscar por Secciones de Tramos
                # ==============================================================
                url_seccion = f"{BASE_URLS['secciones_tramos']}/{eq_id}/"
                async with session.get(url_seccion, headers=HEADERS) as resp:
                    data_seccion = await resp.json() if resp.status == 200 else None
                
                if data_seccion and data_seccion.get('id_tramo'):
                    subestacion = data_seccion.get('linea_nombre') or 'Línea de Transmisión'
                    
                    # Consultamos directamente el endpoint de paños vinculados a esta sección
                    url_p_tramo = f"{BASE_URLS['secciones_tramos']}/{eq_id}/panos/"
                    async with session.get(url_p_tramo, headers=HEADERS) as resp_p:
                        panos_data = await resp_p.json() if resp_p.status == 200 else []
                    
                    if panos_data:
                        for p in panos_data:
                            if p.get('id') and p.get('nemotecnico'):
                                panos_a_procesar.append((p.get('id'), p.get('nemotecnico')))
                    else:
                        # Respaldo homólogo por texto si el endpoint del tramo viene vacío
                        url_tramo = f"{BASE_URLS['tramos']}/{data_seccion['id_tramo']}/"
                        async with session.get(url_tramo, headers=HEADERS) as resp_t:
                            data_tramo = await resp_t.json() if resp_t.status == 200 else None
                        
                        if data_tramo:
                            for ext in [data_tramo.get('extremo1_descripcion', ''), data_tramo.get('extremo2_descripcion', '')]:
                                if ext:
                                    ext_limpio = re.sub(r'^(Paño\s*:\s*|Tap\s*:\s*|S/E\s*)', '', ext, flags=re.IGNORECASE).strip()
                                    url_search_p = f"{BASE_URLS['panos']}?nombre__icontains={ext_limpio}"
                                    async with session.get(url_search_p, headers=HEADERS) as resp_sp:
                                        sp_data = await resp_sp.json() if resp_sp.status == 200 else []
                                    if sp_data:
                                        panos_a_procesar.append((sp_data[0].get('id'), sp_data[0].get('nemotecnico')))
            else:
                # ==============================================================
                # MODO DIRECTO: Buscar por Equipos Directos
                # ==============================================================
                # 1. Intentar como Interruptor
                url_eq = f"{BASE_URLS['interruptores']}/{eq_id}"
                async with session.get(url_eq, headers=HEADERS) as resp:
                    data_eq = await resp.json() if resp.status == 200 else None
                if data_eq:
                    subestacion = data_eq.get('subestacion_nombre', 'Desconocida')
                    if data_eq.get('id_pano') and data_eq.get('pano_nombre'):
                        panos_a_procesar.append((data_eq.get('id_pano'), data_eq.get('pano_nombre')))

                # 2. Intentar como Trafo 2D
                if not panos_a_procesar:
                    url_trafo2d = f"{BASE_URLS['transformadores_2d']}/{eq_id}"
                    async with session.get(url_trafo2d, headers=HEADERS) as resp:
                        data_eq = await resp.json() if resp.status == 200 else None
                    if data_eq:
                        subestacion = data_eq.get('subestacion_nombre', 'Desconocida')
                        if data_eq.get('id_pano') and data_eq.get('pano_nombre'):
                            panos_a_procesar.append((data_eq.get('id_pano'), data_eq.get('pano_nombre')))

                # 3. Intentar como Trafo 3D
                if not panos_a_procesar:
                    url_trafo3d = f"{BASE_URLS['transformadores_3d']}/{eq_id}"
                    async with session.get(url_trafo3d, headers=HEADERS) as resp:
                        data_eq = await resp.json() if resp.status == 200 else None
                    if data_eq:
                        subestacion = data_eq.get('subestacion_nombre', 'Desconocida')
                        if data_eq.get('id_pano') and data_eq.get('pano_nombre'):
                            panos_a_procesar.append((data_eq.get('id_pano'), data_eq.get('pano_nombre')))

            if not panos_a_procesar: continue

            # ==============================================================
            # EXTRACCIÓN SÓLIDA DE COMPONENTES DEL PAÑO ESPECÍFICO
            # ==============================================================
            # Eliminamos duplicados de paños por ID para no repetir búsquedas
            panos_unicos = {p[0]: p[1] for p in panos_a_procesar if p[0]}
            
            for id_pano, nemotecnico_pano in panos_unicos.items():
                endpoints_series = ['interruptores', 'desconectadores', 'transformadores_corriente', 'trampas_ondas']
                sub_equipos_encontrados = []

                for tipo in endpoints_series:
                    # Filtramos de forma ultra-precisa usando el ID del paño directo en la API (?id_pano=)
                    url_search = f"{BASE_URLS[tipo]}/?id_pano={id_pano}"
                    async with session.get(url_search, headers=HEADERS) as resp:
                        datos_api = await resp.json() if resp.status == 200 else []
                    
                    # Si la API no soporta id_pano para ese tipo, caemos al filtro estricto por nemotécnico exacto
                    if not datos_api and nemotecnico_pano:
                        url_search = f"{BASE_URLS[tipo]}/?search={nemotecnico_pano}"
                        async with session.get(url_search, headers=HEADERS) as resp:
                            brutos = await resp.json() if resp.status == 200 else []
                        # Forzamos a que el código del paño esté presente sí o sí en el nemotécnico del equipo
                        datos_api = [x for x in brutos if nemotecnico_pano.lower() in x.get('nombre', '').lower() or nemotecnico_pano.lower() in x.get('nemotecnico', '').lower()]

                    for item in datos_api:
                        url_ficha = f"{BASE_URLS[tipo]}/{item['id']}/fichas-tecnicas/general/"
                        async with session.get(url_ficha, headers=HEADERS) as resp:
                            ficha = await resp.json() if resp.status == 200 else None
                        
                        if ficha:
                            id_campo_corr = '6019' if tipo == 'interruptores' else '6216' if tipo == 'desconectadores' else '6177' if tipo == 'transformadores_corriente' else '469'
                            txt_corr = ficha.get(id_campo_corr, {}).get('valor_texto', '')
                            valor_amp = limpiar_valor_float(txt_corr)
                            
                            valor_ruptura = float('inf')
                            if tipo == 'interruptores':
                                txt_rup = ficha.get('326', {}).get('valor_texto', '')
                                valor_ruptura = limpiar_valor_float(txt_rup)

                            if valor_amp != float('inf') or valor_ruptura != float('inf'):
                                sub_equipos_encontrados.append({
                                    'id': item['id'],
                                    'nombre': item.get('nombre', f"{tipo}_{item['id']}"),
                                    'tipo': tipo.replace("_", " ").upper(),
                                    'corriente': valor_amp,
                                    'ruptura': valor_ruptura
                                })

                if sub_equipos_encontrados:
                    max_equipos_detectados = max(max_equipos_detectados, len(sub_equipos_encontrados))
                    
                    limitante_corriente = min(sub_equipos_encontrados, key=lambda x: x['corriente'])
                    equipos_con_ruptura = [x for x in sub_equipos_encontrados if x['ruptura'] != float('inf')]
                    limitante_ruptura = min(equipos_con_ruptura, key=lambda x: x['ruptura']) if equipos_con_ruptura else None

                    filas_reporte.append({
                        'id_consultado': eq_id,
                        'subestacion': subestacion,
                        'pano_nombre': nemotecnico_pano,
                        'equipos': sub_equipos_encontrados,
                        'lim_corriente': f"{limitante_corriente['nombre']} (ID: {limitante_corriente['id']}) - {limitante_corriente['corriente']} A",
                        'lim_ruptura': f"{limitante_ruptura['nombre']} (ID: {limitante_ruptura['id']}) - {limitante_ruptura['ruptura']} kA" if limitante_ruptura else "N/A"
                    })

        if not filas_reporte:
            raise ValueError("No se encontraron elementos en serie. Asegúrate de ingresar los IDs correctos para el modo seleccionado (Directo para Equipos, Tramo para Líneas).")

        # ==============================================================
        # ESCRITURA EN FORMATO TOTALMENTE HORIZONTAL
        # ==============================================================
        headers = ['ID Consultado', 'Subestación / Elemento', 'Paño Coordinado']
        for i in range(1, max_equipos_detectados + 1):
            headers.extend([
                f'Equipo {i} ID', f'Equipo {i} Nombre', f'Equipo {i} Tipo', f'Equipo {i} Corr [A]', f'Equipo {i} Rup [kA]'
            ])
        headers.extend(['Elemento Limitante Corriente', 'Elemento Limitante Ruptura'])
        ws.append(headers)
        
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        block_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        for r_data in filas_reporte:
            fila_completa = [r_data['id_consultado'], r_data['subestacion'], r_data['pano_nombre']]
            
            for eq in r_data['equipos']:
                corr_display = eq['corriente'] if eq['corriente'] != float('inf') else 'N/A'
                rup_display = eq['ruptura'] if eq['ruptura'] != float('inf') else 'N/A'
                fila_completa.extend([eq['id'], eq['nombre'], eq['tipo'], corr_display, rup_display])
                
            celdas_faltantes = (max_equipos_detectados - len(r_data['equipos'])) * 5
            if celdas_faltantes > 0:
                fila_completa.extend([""] * celdas_faltantes)
                
            fila_completa.extend([r_data['lim_corriente'], r_data['lim_ruptura']])
            ws.append(fila_completa)
            
            row_idx = ws.max_row
            total_cols = len(headers)
            
            celda_m_corr = ws.cell(row=row_idx, column=total_cols - 1)
            celda_m_rup = ws.cell(row=row_idx, column=total_cols)
            
            for c_res in [celda_m_corr, celda_m_rup]:
                c_res.fill = block_fill
                c_res.font = Font(name='Arial', size=10, bold=True, color='C00000')
                
            for col_idx in range(1, total_cols + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical='center', horizontal='center')

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        filepath = os.path.abspath(os.path.join(os.getcwd(), "equipos_datos.xlsx"))
        wb.save(filepath)
        return filepath
